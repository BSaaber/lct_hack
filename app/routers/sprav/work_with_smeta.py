import re

from openpyxl import load_workbook
from io import BytesIO
from enum import Enum
from . import schemas
from fastapi import HTTPException, status
from collections import namedtuple
import app.database.api as db_api
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Union


class ColumnIndexes(BaseModel):
    code: Union[int, None] = None
    name: Union[int, None] = None
    uom: Union[int, None] = None
    amount: Union[int, None] = None
    price: Union[int, None] = None


class ParseMode(Enum):
    FINDING_ADDRESS = 1
    FINDING_COLUMNS_INDEXES = 2
    PARSING_LINES = 3


class SmetaLineStandard(Enum):
    TSN = 1
    SN = 2
    UNDEFINED = 3


sn_pattern = r'\d+\.\d+-\d+-\d+-\d+/\d+'
#              5   .4  -320-7  -1  /1     | 5.4-320-7-1/1

tsn_pattern = r'\d+.\d+-\d+-\d+'


#               3  .51 -2  -1             | 3.51-2-1


def get_smeta_standard(code):
    if re.fullmatch(sn_pattern, code):
        return SmetaLineStandard.SN
    elif re.fullmatch(tsn_pattern, code):
        return SmetaLineStandard.TSN
    else:
        return SmetaLineStandard.UNDEFINED


def cell_not_empty(cell):
    return isinstance(cell, str) and cell != ''


def cell_is_empty(cell):
    return not isinstance(cell, str) or cell == ''


# TODO что делать с хуетой по типу Source!=F28

async def parse_smeta(db: Session, file: bytes):



    result = schemas.Smeta()
    smeta_category = schemas.SmetaCategory()
    smeta_line = schemas.SmetaLine()
    smeta_hypothesis = schemas.SpgzHypothesis()
    for i in range(5):
        smeta_line.hypothesises.append(smeta_hypothesis)
    for i in range(15):
        smeta_category.lines.append(smeta_line)
    for i in range(3):
        result.categories.append(smeta_category)
    return result








    workbook = load_workbook(BytesIO(file), read_only=True)  # .read()?
    worksheet = workbook.active
    result = schemas.Smeta()
    mode = ParseMode.FINDING_ADDRESS
    column_indexes = ColumnIndexes()
    for line_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if mode is ParseMode.FINDING_ADDRESS:
            mode = ParseMode.FINDING_COLUMNS_INDEXES
        elif mode is ParseMode.FINDING_COLUMNS_INDEXES:
            print(line_index, "| ", row)
            for j, val in enumerate(row):
                if line_index == 26:
                    print(f"val {j}: {val}")
                if cell_not_empty(val) and line_index == 26:
                    print("val is not empty")
                    if 'шифр' in val.lower():
                        column_indexes.code = j
                    elif "наименование" in val.lower():
                        column_indexes.name = j
                    elif "ед. изм." in val.lower() or "измерен" in val.lower():
                        column_indexes.uom = j
                    elif "кол-во" in val.lower() or "количество" in val.lower():
                        column_indexes.amount = j
                    elif "всего" in val.lower():
                        column_indexes.price = j
            print("now column_indexes is: ", column_indexes)
            print(column_indexes.dict().values())
            print("---")
            if None not in column_indexes.dict().values():
                mode = ParseMode.PARSING_LINES
                print("MODE CHANGED TO PARSING LINES")
                print("------------------------------")
            # not_found = ', '.join([k for k, v in column_indexes._asdict().items() if v is None])
            # if not_found != '':
            # raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
            #                    detail=f"error: columns {not_found} are not found")
        elif mode is ParseMode.PARSING_LINES:
            if cell_not_empty(row[0]) and 'раздел' in row[0].lower():  # new category
                print("STARTING NEW CATEGORY")
                result.categories.append(schemas.SmetaCategory())
                continue
            elif cell_not_empty(row[column_indexes.code]) and cell_is_empty(row[column_indexes.price]):  # new line
                print("NEW LINE")
                result.categories[-1].lines.append(schemas.SmetaLine())
                result.categories[-1].lines[-1].code = row[column_indexes.code]
                result.categories[-1].lines[-1].name = row[column_indexes.name]
                result.categories[-1].lines[-1].uom = row[column_indexes.uom]
                result.categories[-1].lines[-1].amount = row[column_indexes.amount]
                result.categories[-1].lines[-1].code = row[column_indexes.code]
                result.categories[-1].lines[-1].line_number = line_index + 1
            elif cell_is_empty(row[column_indexes.code]) and cell_not_empty(row[column_indexes.price]):  # price
                print("NEW LINE PRICE")
                result.categories[-1].lines[-1].price = row[column_indexes.price]

                standard = get_smeta_standard(result.categories[-1].lines[-1].code)
                if standard == SmetaLineStandard.SN:
                    sn_piece = await db_api.sprav_edit.get_tsn_piece_by_code(db, result.categories[-1].lines[-1].code)
                    if sn_piece is None:
                        raise HTTPException(status_code=status.HTTP_500_BAD_REQUEST,
                                            detail=f"error: no data for code in line {line_index}: {result.categories[-1].lines[-1].code}\ncurrent building line: {result.categories[-1].lines[-1]}")
                    hypothesises = await db_api.sprav_edit.get_sn_hypothesises_by_tsn_id(db, sn_piece.id)
                    if hypothesises is None:
                        raise HTTPException(status_code=status.HTTP_500_BAD_REQUEST,
                                            detail=f"error: no hypothesises for code in line {line_index}: {result.categories[-1].lines[-1].code}\ncurrent building line: {result.categories[-1].lines[-1]}")
                    result.categories[-1].lines[-1].hypothesises = hypothesises
                elif standard == SmetaLineStandard.TSN:
                    pass
                elif standard == SmetaLineStandard.UNDEFINED:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                        detail=f"bad code format in line {line_index}: {result.categories[-1].lines[-1].code}\ncurrent building line: {result.categories[-1].lines[-1]}")
            else:  # case trash
                pass

    return result
