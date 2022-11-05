from openpyxl import load_workbook
from sqlalchemy import create_engine
from app.database import db_models  # noqa - for db initialization
from app.database import schemas as db_schemas
import app.database.api as db_api
import os
from dotenv import load_dotenv
import psycopg2  # noqa - driver for db
from sqlalchemy.orm import sessionmaker
import asyncio

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, ".env"))

engine = create_engine(os.environ["DATABASE_URL"])
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


# TODO check read only mode in load_workbook [  mean load_workbook(filename, read_only=True) ]


async def parse_sn(filename: str):
    print("parse_sn started to work")
    with SessionLocal() as db:
        workbook = load_workbook(filename, read_only=True)
        worksheet = workbook.active
        errors = 0
        code = None
        text = None
        uom = None
        price = None
        skipper = 0
        insert = False
        waiter = False
        dealt = 0
        check_for_end = False
        code_index = 0
        text_index = 0
        uom_index = 0
        price_index = 0
        for row in worksheet.iter_rows(min_row=9, max_row=9, values_only=True):
            # print(row)
            for i, cell in enumerate(row):
                if cell == "Шифр, номера нормативов и коды ресурсов":
                    code_index = i
                elif cell == "Наименование работ и затрат":
                    text_index = i
                elif cell == "Ед. изм.":
                    uom_index = i
                elif cell == "ВСЕГО затрат, руб.":
                    price_index = i
        for i, row in enumerate(worksheet.iter_rows(min_row=16, values_only=True)):
            if row[4] == "Итого по разделу":
                print("ended at line: ", i + 28)
                print("ALL WENT DONE")
                break
            if check_for_end:
                print("ERROR!")
                break
            if skipper > 0:
                skipper -= 1
                continue
            if waiter:
                if (row[text_index] is None or row[text_index] == '') and (
                        row[price_index] is not None and row[price_index] != ''):
                    waiter = False
                else:
                    continue

            # print(row)

            if row[0] is not None and row[0] != '':
                code = row[code_index]
                text = row[text_index]
                uom = row[uom_index]
                if row[price_index] is not None and row[price_index] != '':
                    price = row[price_index]
                    skipper = 2
                    # print("determine skipper := 2")
                else:
                    waiter = True
                    # print("determine waiter := True")
                    continue
            else:
                price = row[price_index]
                skipper = 1
                # print("determine skipper := 1")

            # print("line: ", i + 28)
            # print("code: ", code)
            # print("text: ", text)
            # print("uom: ", uom)
            # print("price: ", price)
            if code is None or text is None or uom is None or price is None:
                print("line: ", i + 28)
                print("code: ", code)
                print("text: ", text)
                print("uom: ", uom)
                print("price: ", price)
                check_for_end = True
                continue
            else:
                dealt += 1
            # code = None
            # text = None
            # uom = None
            # price = None
            # print("----------------")
            # form sn_piece

            sn_piece = db_schemas.SnPieceCreateWithoutSpgz(code=code, text=text, uom=uom, price=price)
            sn_result = await db_api.sprav_edit.get_sn_piece_by_code(db, code)
            if sn_result is None:
                # write sn_piece to db
                sn_result = await db_api.sprav_edit.add_sn_piece_without_spgz(db, sn_piece)
                if sn_result is None:
                    errors += 1
                    print("error during writing spgz to db")
                    print(row)
        return errors, dealt


filenames = os.listdir("app/scripts/data/sn")
total_errors = 0
loop = asyncio.get_event_loop()
total_dealt = 0
for dir_filename in filenames:
    print("working with ", dir_filename)
    errors_amount, dealt_amount = loop.run_until_complete(parse_sn("app/scripts/data/sn/" + dir_filename))
    print("errors: ")
    print(errors_amount)
    print("dealt: ")
    print(dealt_amount)
    total_dealt += dealt_amount
    total_errors += errors_amount
loop.close()

print("total errors: ")
print(total_errors)
print("total dealt: ")
print(total_dealt)
