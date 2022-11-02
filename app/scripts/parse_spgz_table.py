from openpyxl import load_workbook
from sqlalchemy import create_engine
from app.database import db_models # noqa - for db initialization
from app.database import schemas as db_schemas
import app.database.api as db_api
import psycopg2

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, ".env"))

engine = create_engine(os.environ["DATABASE_URL"])
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def parse_spgz_kpgz(filename: str):
    workbook = load_workbook(filename)
    worksheet = workbook.active
    errors = 0
    for row in worksheet.iter_rows(min_row=2, max_col=9, values_only=True):
        data_id = row[0]
        kpgz_name = row[3]
        spgz_name = row[4]
        spgz_description = row[5]
        uom = row[6]
        okpd = row[7]
        okpd2 = row[8]
        kpgz_piece = db_schemas.KpgzPieceCreate(name=kpgz_name)
        kpgz_result = db_api.sprav_edit.add_kpgz_piece(get_db(), kpgz_piece)
        if kpgz_result is None:
            errors += 1
            print(row)

        spgz_piece = db_schemas.SpgzPieceCreate(name=spgz_name, okpd=okpd, okpd2=okpd2, uom=uom,description=spgz_description, data_id=data_id, kpgz_piece_id=kpgz_result.id)

        db_api.sprav_edit.add_spgz_piece(get_db(), spgz_piece)
    return errors


errors_amount = parse_tsn("spgz_kpgz.xlsx")
print("total errors: ")
print(errors_amount)
