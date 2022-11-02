from openpyxl import load_workbook


def parse_tsn(filename: str):
    workbook = load_workbook(filename)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, max_col=9, values_only=True):
      pass


parse_tsn("tsn1.xlsx")
