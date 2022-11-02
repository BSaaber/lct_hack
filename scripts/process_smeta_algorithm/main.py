from openpyxl import load_workbook


def parse_tsn(filename: str):
    workbook = load_workbook(filename)
    worksheet = workbook.active
    row_index = 28
    result = []
    while row_index < 83:
        t = {}
        t["first"] = worksheet.cell(row_index, 1).value
        t["code"] = worksheet.cell(row_index, 2).value
        t["text"] = worksheet.cell(row_index, 5).value
        result.append(t)
        row_index += 1
        while worksheet.cell(row_index, 1).value is None:
            row_index += 1
        t["summ"] = worksheet.cell(row_index - 2, 27).value

    for i in result:
        print(i)


parse_tsn("tsn1.xlsx")
